import json
import os
import xlwt

from bank_manager.BankDatas import AccountDatas
import openpyxl

from config import base_path


class FileTool:
    def __init__(self, file_name):
        file_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.dir = os.path.join(file_path, "data", file_name)
        self.workbook = openpyxl.Workbook()
        # self.worksheet = self.workbook.create_sheet('data')
        self.worksheet = self.workbook["Sheet"]

    # 把BankDatas里的数据，写入到excel里
    # 把json格式的数据Data写入到指定路径下文件名为filename的文件中
    def write_json_to_excel(self, data, filename):
        # file_path =os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        # self.dir = os.path.join(file_path, "data",filename)
        # self.workbook = xlwt.Workbook(encoding='utf-8')
        # self.worksheet = self.workbook.add_sheet('data')
        # workbook = openpyxl.load_workbook('data')
        # worksheet = workbook.create_sheet(title='file_data',index=0)

        # 写表头
        for index, val in enumerate(data[0].keys()):
            self.worksheet.cell(1, index + 1).value = val

        # 写数据
        for row, list_item in enumerate(data):
            row += 2
            col = 1
            for key, value in list_item.items():
                self.worksheet.cell(row, col).value = value
                col += 1

        # 保存
        self.workbook.save(self.dir)

    # 把excel的数据读取出来
    def write_data(self, x_y, pwd):
        try:
            self.worksheet.cell(x_y[0] + 1, x_y[1] + 1).value = pwd
        except Exception as e:
            self.worksheet.cell(x_y[0] + 1, x_y[1] + 1).value = e
        finally:
            self.workbook.save(self.dir)

    def read_excel_data(self, x_y):
        return self.worksheet.cell(x_y[0] + 1, x_y[1] + 1).value

    def read_excel_all_data(self):
        case_list = []
        datas = list(self.worksheet.values)
        for value in datas[1::]:
            dict_data = dict(zip(datas[0],value))
            case_list.append(dict_data)
        return case_list


if __name__ == '__main__':
    # 1.把data读到json文档，得到一个json文档
    # FileTool().get_json("case.json", AccountDatas)
    file_tool = FileTool("cases.xlsx")
    file_tool.write_json_to_excel(AccountDatas, "cases.xlsx")
    file_tool.write_data([1, 10], "123456")
